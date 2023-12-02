import cn2an
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

from service import dpms_service


def get_sheet_titles():
    return ("编号",
            "姓名",
            "性别",
            "电话",
            "状态",
            "医院",
            "科室",
            "医生",
            "预购批次",
            "预购时间",
            "预购数量",
            "实际购买时间",
            "购买数量",
            "赠送/直购",
            "备注")


def get_patient_sheet_props():
    return 'id', 'name', 'sex', 'phone', 'status', 'hospital', 'department', 'doctors'


def get_expected_sheet_props():
    return 'expected_at', 'dose'


def get_record_sheet_props():
    return 'record_at', 'dose', 'status', 'remark'


def get_expected_batch(index):
    return "第{}批".format(cn2an.an2cn(index))


def set_sheet_title(sheet):
    for index, title in enumerate(get_sheet_titles()):
        sheet.cell(1, index + 1, title)


def get_solid_color(index):
    if index % 2 == 0:
        return PatternFill("solid", fgColor="BEBEBE")
    return PatternFill("solid", fgColor="FFFFE0")


def get_expected_solid_color(index):
    if index % 2 == 0:
        return PatternFill("solid", fgColor="B4EEB4")
    return PatternFill("solid", fgColor="F0FFF0")


def get_record_solid_color(index):
    if index % 2 == 0:
        return PatternFill("solid", fgColor="BEBEBE")
    return PatternFill("solid", fgColor="FFFFE0")


def get_patient_status_color(status):
    if status == 0:
        return PatternFill("solid", fgColor="FF6A6A")
    if status == 1:
        return PatternFill("solid", fgColor="00FF7F")
    return PatternFill("solid", fgColor="FFD700")


def get_align():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def get_sex(sex):
    if not sex:
        return "保密"
    if 0 == sex:
        return "女"
    if 1 == sex:
        return "男"


class OutPutExcel:

    def __init__(self):
        book = Workbook()
        sheet = book.active
        set_sheet_title(sheet)
        self.set_pdms_data(sheet)
        book.save('data/test.xlsx')

    def set_pdms_data(self, sheet):
        pdms_data = dpms_service.get_excel_filter_data()
        row_start = 2
        # 注意！！index 是从 0 开始，写入Excel需要+2
        for index, dpm in enumerate(pdms_data):
            if index == 0:
                expected_len = 1
            else:
                expected_len = max(len(pdms_data[index-1].get('patient_expected')), len(pdms_data[index-1].get('patient_record')))
            if expected_len < 1:
                expected_len = 1
            # 录入患者基本信息
            for colum, patient_sheet in enumerate(get_patient_sheet_props()):
                sheet_string = dpm.get(patient_sheet)
                # if colum == 2:
                #     sheet_string = get_sex(sheet_string)
                self.set_patient_sheet(sheet, index, row_start, colum + 1, sheet_string, expected_len)
            # 录入患者预购信息
            self.set_expected_data(sheet, dpm.get('patient_expected'), row_start, 10, index)
            # 录入患者已购信息
            self.set_record_data(sheet, dpm.get('patient_record'), row_start, 12, index)
            row_start += expected_len

    # 患者预购信息写入
    def set_expected_data(self, sheet, expectedDatas, row_start, colum_start, index):
        for row, expect in enumerate(expectedDatas):
            sheet.cell(row_start + row, 9, get_expected_batch(row+1))
            sheet.cell(row_start + row, 9).fill = get_expected_solid_color(index)

            for colum, prop in enumerate(get_expected_sheet_props()):
                sheet.cell(row_start + row, colum_start + colum, expect.get(prop))
                sheet.cell(row_start + row, colum_start + colum).fill = get_expected_solid_color(index)

    # 患者购药信息写入
    def set_record_data(self, sheet, recordDatas, row_start, colum_start, index):
        for row, expect in enumerate(recordDatas):
            for colum, prop in enumerate(get_record_sheet_props()):
                sheet.cell(row_start + row, colum_start + colum, expect.get(prop))
                sheet.cell(row_start + row, colum_start + colum).fill = get_record_solid_color(index)

    # 患者基本信息写入 Excel
    def set_patient_sheet(self, sheet, index, start_row, start_col, cell_text, expected_len=1):
        sheet.cell(start_row, start_col, cell_text)
        sheet.cell(start_row, start_col).alignment = get_align()
        sheet.cell(start_row, start_col).fill = get_solid_color(index)
        sheet.merge_cells(start_row=start_row,
                          end_row=start_row + expected_len - 1,
                          start_column=start_col,
                          end_column=start_col)

